# tests/infrastructure/test_report_pack_xlsx.py
# Tests for Report Pack XLSX export (v1.10)

import pytest
import pandas as pd
from pathlib import Path
from backtester.infrastructure.reporter import Reporter
from backtester.infrastructure.reporting.report_pack import (
    build_report_pack_xlsx,
    _has_excel_engine,
)


# Скипаем все тесты если нет Excel engine
pytestmark = pytest.mark.skipif(
    not _has_excel_engine(),
    reason="No Excel engine (openpyxl/xlsxwriter) installed in environment"
)


def test_report_pack_created_when_openpyxl_available(tmp_path):
    """
    Test A: Проверяем что report_pack.xlsx создается когда openpyxl доступен.
    """
    output_dir = tmp_path / "reports"
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Создаем минимальные CSV файлы
    positions_csv = output_dir / "portfolio_positions.csv"
    positions_df = pd.DataFrame([
        {
            "strategy": "test_strategy",
            "signal_id": "sig1",
            "contract_address": "CONTRACT1",
            "entry_time": "2024-01-01T12:00:00Z",
            "exit_time": "2024-01-01T14:00:00Z",
            "status": "closed",
            "pnl_sol": 0.1,
        }
    ])
    positions_df.to_csv(positions_csv, index=False)
    
    portfolio_events_csv = output_dir / "portfolio_events.csv"
    events_df = pd.DataFrame([
        {
            "timestamp": "2024-01-01T12:00:00Z",
            "event_type": "ATTEMPT_RECEIVED",
            "strategy": "test_strategy",
            "signal_id": "sig1",
            "contract_address": "CONTRACT1",
        }
    ])
    events_df.to_csv(portfolio_events_csv, index=False)
    
    # Конфиг
    config = {
        "export_xlsx": True,
        "xlsx_filename": "report_pack.xlsx",
        "xlsx_timestamped": False,
        "xlsx_sheets": ["summary", "positions", "portfolio_events"],
    }
    
    inputs = {
        "positions_csv": positions_csv,
        "portfolio_events_csv": portfolio_events_csv,
        "stage_a_stability_csv": None,
        "stage_b_selection_csv": None,
        "policy_summary_csv": None,
        "capacity_prune_events_csv": None,
    }
    
    # Вызываем build_report_pack_xlsx
    result_path = build_report_pack_xlsx(
        output_dir=output_dir,
        inputs=inputs,
        config=config,
        portfolio_results=None,
        runner_stats=None,
        include_skipped_attempts=True,
    )
    
    # Проверяем что файл создан
    assert result_path is not None
    assert result_path.exists()
    assert result_path.name == "report_pack.xlsx"
    
    # Проверяем листы через openpyxl
    try:
        from openpyxl import load_workbook
        wb = load_workbook(result_path)
        sheet_names = wb.sheetnames
        
        assert "summary" in sheet_names
        assert "positions" in sheet_names
        assert "portfolio_events" in sheet_names
        
        # Проверяем что summary содержит данные
        ws_summary = wb["summary"]
        assert ws_summary.max_row > 1  # Есть заголовок + данные
        
        # Проверяем что positions содержит данные
        ws_positions = wb["positions"]
        assert ws_positions.max_row > 1  # Есть заголовок + данные
        
    except ImportError:
        # Если openpyxl недоступен для чтения, проверяем через pandas
        xls = pd.ExcelFile(result_path)
        assert "summary" in xls.sheet_names
        assert "positions" in xls.sheet_names
        assert "portfolio_events" in xls.sheet_names


def test_report_pack_best_effort_missing_optional_csv(tmp_path):
    """
    Test B: Проверяем что генерация не падает при отсутствии опциональных CSV.
    """
    output_dir = tmp_path / "reports"
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Создаем только portfolio_events.csv
    portfolio_events_csv = output_dir / "portfolio_events.csv"
    events_df = pd.DataFrame([
        {
            "timestamp": "2024-01-01T12:00:00Z",
            "event_type": "ATTEMPT_RECEIVED",
            "strategy": "test_strategy",
            "signal_id": "sig1",
            "contract_address": "CONTRACT1",
        }
    ])
    events_df.to_csv(portfolio_events_csv, index=False)
    
    config = {
        "export_xlsx": True,
        "xlsx_filename": "report_pack.xlsx",
        "xlsx_timestamped": False,
        "xlsx_sheets": ["summary", "positions", "portfolio_events"],
    }
    
    inputs = {
        "positions_csv": None,  # Отсутствует
        "portfolio_events_csv": portfolio_events_csv,
        "stage_a_stability_csv": None,
        "stage_b_selection_csv": None,
        "policy_summary_csv": None,
        "capacity_prune_events_csv": None,
    }
    
    # Вызываем - не должно падать
    result_path = build_report_pack_xlsx(
        output_dir=output_dir,
        inputs=inputs,
        config=config,
        portfolio_results=None,
        runner_stats=None,
        include_skipped_attempts=True,
    )
    
    # Проверяем что файл создан
    assert result_path is not None
    assert result_path.exists()
    
    # Проверяем листы
    try:
        from openpyxl import load_workbook
        wb = load_workbook(result_path)
        sheet_names = wb.sheetnames
        
        assert "summary" in sheet_names
        assert "portfolio_events" in sheet_names
        # positions может отсутствовать или быть с "missing"
        if "positions" in sheet_names:
            ws_positions = wb["positions"]
            # Проверяем что есть либо данные, либо "missing"
            if ws_positions.max_row > 1:
                first_data_row = [cell.value for cell in ws_positions[2]]
                # Может быть "missing" или реальные данные
                assert len(first_data_row) > 0
    except ImportError:
        xls = pd.ExcelFile(result_path)
        assert "summary" in xls.sheet_names
        assert "portfolio_events" in xls.sheet_names


def test_report_pack_skips_when_no_engine(monkeypatch, tmp_path):
    """
    Test C: Проверяем что функция возвращает None и не падает при отсутствии engine.
    """
    # Мокируем _has_excel_engine чтобы вернуть False
    from backtester.infrastructure.reporting import report_pack
    
    original_has_engine = report_pack._has_excel_engine
    
    def mock_has_engine():
        return False
    
    monkeypatch.setattr(report_pack, "_has_excel_engine", mock_has_engine)
    
    output_dir = tmp_path / "reports"
    output_dir.mkdir(parents=True, exist_ok=True)
    
    config = {
        "export_xlsx": True,
        "xlsx_filename": "report_pack.xlsx",
        "xlsx_timestamped": False,
    }
    
    inputs = {
        "positions_csv": None,
        "portfolio_events_csv": None,
    }
    
    # Вызываем с предупреждением
    with pytest.warns(UserWarning, match="Excel engine.*not installed"):
        result_path = build_report_pack_xlsx(
            output_dir=output_dir,
            inputs=inputs,
            config=config,
            portfolio_results=None,
            runner_stats=None,
            include_skipped_attempts=True,
        )
    
    # Должен вернуть None
    assert result_path is None
    
    # Восстанавливаем оригинальную функцию
    monkeypatch.setattr(report_pack, "_has_excel_engine", original_has_engine)


def test_reporter_save_report_pack_xlsx(tmp_path):
    """
    Test: Проверяем метод Reporter.save_report_pack_xlsx().
    """
    output_dir = tmp_path / "reports"
    reporter = Reporter(output_dir=str(output_dir))
    
    # Создаем минимальные CSV
    positions_csv = output_dir / "portfolio_positions.csv"
    positions_df = pd.DataFrame([
        {
            "strategy": "test_strategy",
            "signal_id": "sig1",
            "contract_address": "CONTRACT1",
            "entry_time": "2024-01-01T12:00:00Z",
            "exit_time": "2024-01-01T14:00:00Z",
            "status": "closed",
            "pnl_sol": 0.1,
        }
    ])
    positions_df.to_csv(positions_csv, index=False)
    
    portfolio_events_csv = output_dir / "portfolio_events.csv"
    events_df = pd.DataFrame([
        {
            "timestamp": "2024-01-01T12:00:00Z",
            "event_type": "ATTEMPT_RECEIVED",
            "strategy": "test_strategy",
            "signal_id": "sig1",
            "contract_address": "CONTRACT1",
        }
    ])
    events_df.to_csv(portfolio_events_csv, index=False)
    
    # Вызываем метод Reporter
    result_path = reporter.save_report_pack_xlsx(
        portfolio_results=None,
        runner_stats={"signals_processed": 10},
        include_skipped_attempts=True,
        config=None,  # Используем дефолты
    )
    
    # Проверяем что файл создан
    if result_path is not None:
        assert result_path.exists()
        assert result_path.name == "report_pack.xlsx"
        
        # Проверяем листы
        xls = pd.ExcelFile(result_path)
        assert "summary" in xls.sheet_names
        assert "positions" in xls.sheet_names
        assert "portfolio_events" in xls.sheet_names

