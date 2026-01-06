# backtester/infrastructure/reporting/report_pack.py
# v1.10: Report Pack - единый XLSX-отчёт со всеми ключевыми таблицами

from __future__ import annotations

import warnings
import csv
from pathlib import Path
from typing import Dict, Optional, List
from datetime import datetime, timezone


def _has_excel_engine() -> bool:
    """
    Проверяет наличие доступного Excel engine.
    
    Returns:
        True если есть openpyxl или xlsxwriter, иначе False
    """
    try:
        import openpyxl  # noqa
        return True
    except ImportError:
        pass
    try:
        import xlsxwriter  # noqa
        return True
    except ImportError:
        return False


def _pick_excel_engine() -> str:
    """
    Выбирает engine для Excel writer с fallback.
    
    Приоритет:
    1. openpyxl (предпочтительно для report_pack)
    2. xlsxwriter (fallback)
    
    Raises:
        ImportError если ни один движок не установлен
    """
    try:
        import openpyxl  # noqa
        return "openpyxl"
    except ImportError:
        try:
            import xlsxwriter  # noqa
            return "xlsxwriter"
        except ImportError:
            raise ImportError("Neither openpyxl nor xlsxwriter is installed")


def _read_csv_to_rows(csv_path: Path) -> Optional[List[Dict[str, str]]]:
    """
    Читает CSV файл и возвращает список словарей (строки).
    
    Returns:
        Список словарей или None если файл не найден/ошибка чтения
    """
    if not csv_path.exists():
        return None
    
    try:
        rows = []
        with csv_path.open("r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(row)
        return rows
    except Exception as e:
        warnings.warn(
            f"Failed to read CSV {csv_path}: {e}. Skipping sheet.",
            UserWarning,
            stacklevel=3
        )
        return None


def _create_summary_sheet(
    output_dir: Path,
    portfolio_results: Optional[Dict] = None,
    runner_stats: Optional[Dict] = None,
    include_skipped_attempts: bool = True,
) -> List[Dict[str, str]]:
    """
    Создает лист summary с метаданными и топлайновыми метриками.
    
    Returns:
        Список словарей для записи в Excel (key/value пары)
    """
    rows = []
    
    # Блок A: Run metadata
    rows.append({"key": "=== Run Metadata ===", "value": ""})
    rows.append({"key": "run_timestamp_utc", "value": datetime.now(timezone.utc).isoformat()})
    rows.append({"key": "project_version", "value": "v2.1.9"})
    
    # Git commit (если доступно)
    try:
        import subprocess
        result = subprocess.run(
            ["git", "rev-parse", "HEAD"],
            capture_output=True,
            text=True,
            timeout=2
        )
        if result.returncode == 0:
            rows.append({"key": "git_commit", "value": result.stdout.strip()})
        else:
            rows.append({"key": "git_commit", "value": ""})
    except Exception:
        rows.append({"key": "git_commit", "value": ""})
    
    rows.append({"key": "strategy_mode", "value": "runner-only"})
    rows.append({"key": "include_skipped_attempts", "value": str(include_skipped_attempts)})
    rows.append({"key": "", "value": ""})  # Пустая строка
    
    # Блок B: Topline metrics
    rows.append({"key": "=== Topline Metrics ===", "value": ""})
    
    if portfolio_results:
        # Берем первую стратегию для метрик (можно агрегировать, но для summary достаточно первой)
        first_result = next(iter(portfolio_results.values()))
        if hasattr(first_result, "stats"):
            stats = first_result.stats
            rows.append({
                "key": "final_balance_sol",
                "value": str(getattr(stats, "final_balance_sol", 0.0))
            })
            rows.append({
                "key": "total_return_pct",
                "value": str(getattr(stats, "total_return_pct", 0.0))
            })
            rows.append({
                "key": "max_drawdown_pct",
                "value": str(getattr(stats, "max_drawdown_pct", 0.0))
            })
            rows.append({
                "key": "trades_executed",
                "value": str(getattr(stats, "trades_executed", 0))
            })

            # Считаем события из portfolio_events если есть
            # v2.0.1: ATTEMPT_* events removed (Runner-only canonical events only)
            # Legacy attempt tracking removed - no longer tracked in canonical event ledger

            rows.append({
                "key": "portfolio_capacity_prune_count",
                "value": str(getattr(stats, "portfolio_capacity_prune_count", 0))
            })
            rows.append({
                "key": "portfolio_reset_count",
                "value": str(getattr(stats, "portfolio_reset_count", 0))
            })
    
    if runner_stats:
        rows.append({
            "key": "signals_processed",
            "value": str(runner_stats.get("signals_processed", 0))
        })
        rows.append({
            "key": "signals_skipped_no_candles",
            "value": str(runner_stats.get("signals_skipped_no_candles", 0))
        })
        rows.append({
            "key": "signals_skipped_corrupt_candles",
            "value": str(runner_stats.get("signals_skipped_corrupt_candles", 0))
        })
    
    return rows


def build_report_pack_xlsx(
    output_dir: Path,
    inputs: Dict[str, Optional[Path]],
    config: Dict,
    portfolio_results: Optional[Dict] = None,
    runner_stats: Optional[Dict] = None,
    include_skipped_attempts: bool = True,
) -> Optional[Path]:
    """
    Собирает единый XLSX-отчёт из CSV файлов.
    
    Args:
        output_dir: Директория для сохранения report_pack.xlsx
        inputs: Словарь {sheet_name: csv_path} с путями к CSV файлам
        config: Конфигурация (xlsx_filename, xlsx_timestamped, xlsx_sheets)
        portfolio_results: Опционально, для summary метрик
        runner_stats: Опционально, для summary метрик
        include_skipped_attempts: Флаг для summary
    
    Returns:
        Path к созданному файлу или None если не удалось создать
    """
    if not _has_excel_engine():
        warnings.warn(
            "Excel engine (openpyxl/xlsxwriter) not installed; skipping report_pack.xlsx export. "
            "CSV files will still be created.",
            UserWarning,
            stacklevel=2
        )
        return None
    
    engine = _pick_excel_engine()
    
    # Определяем имя файла
    base_filename = config.get("xlsx_filename", "report_pack.xlsx")
    if config.get("xlsx_timestamped", False):
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        name_parts = base_filename.rsplit(".", 1)
        if len(name_parts) == 2:
            filename = f"{name_parts[0]}_{timestamp}.{name_parts[1]}"
        else:
            filename = f"{base_filename}_{timestamp}"
    else:
        filename = base_filename
    
    output_path = output_dir / filename
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Список листов для создания
    sheets_to_create = config.get("xlsx_sheets", [
        "summary",
        "positions",
        "portfolio_events",
        "stage_a_stability",
        "stage_b_selection",
        "policy_summary",
        "capacity_prune_events",
    ])
    
    try:
        if engine == "openpyxl":
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            # Удаляем дефолтный лист
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            
            # Лист summary (всегда первый)
            if "summary" in sheets_to_create:
                ws_summary = wb.create_sheet("summary")
                summary_rows = _create_summary_sheet(
                    output_dir,
                    portfolio_results,
                    runner_stats,
                    include_skipped_attempts
                )
                if summary_rows:
                    # Заголовки
                    ws_summary.append(["key", "value"])
                    for row in summary_rows:
                        ws_summary.append([row.get("key", ""), row.get("value", "")])
                    # Автоширина колонок
                    ws_summary.column_dimensions["A"].width = 30
                    ws_summary.column_dimensions["B"].width = 50
            
            # Остальные листы из CSV
            sheet_mapping = {
                "positions": inputs.get("positions_csv"),
                "portfolio_events": inputs.get("portfolio_events_csv"),
                "stage_a_stability": inputs.get("stage_a_stability_csv"),
                "stage_b_selection": inputs.get("stage_b_selection_csv"),
                "policy_summary": inputs.get("policy_summary_csv"),
                "capacity_prune_events": inputs.get("capacity_prune_events_csv"),
            }
            
            for sheet_name in sheets_to_create:
                if sheet_name == "summary":
                    continue  # Уже создан
                
                csv_path = sheet_mapping.get(sheet_name)
                if csv_path is None:
                    continue
                
                rows = _read_csv_to_rows(csv_path)
                if rows is None:
                    # Создаем лист с "missing"
                    ws = wb.create_sheet(sheet_name)
                    ws.append(["status"])
                    ws.append(["missing"])
                    continue
                
                if not rows:
                    # Пустой CSV - создаем лист с заголовками если возможно
                    ws = wb.create_sheet(sheet_name)
                    ws.append(["empty"])
                    continue
                
                # Создаем лист и записываем данные
                ws = wb.create_sheet(sheet_name)
                
                # Заголовки (из первой строки)
                headers = list(rows[0].keys())
                ws.append(headers)
                
                # Данные
                for row in rows:
                    ws.append([row.get(h, "") for h in headers])
                
                # Автоширина колонок (примерно)
                for idx, header in enumerate(headers, 1):
                    col_letter = get_column_letter(idx)
                    ws.column_dimensions[col_letter].width = max(len(str(header)), 15)
            
            wb.save(output_path)
            print(f"📦 Saved report pack to {output_path}")
            return output_path
            
        else:  # xlsxwriter fallback
            import pandas as pd
            # Используем pandas ExcelWriter напрямую для xlsxwriter
            
            sheets = {}
            
            # Summary
            if "summary" in sheets_to_create:
                summary_rows = _create_summary_sheet(
                    output_dir,
                    portfolio_results,
                    runner_stats,
                    include_skipped_attempts
                )
                if summary_rows:
                    sheets["summary"] = pd.DataFrame(summary_rows)
            
            # Остальные листы
            sheet_mapping = {
                "positions": inputs.get("positions_csv"),
                "portfolio_events": inputs.get("portfolio_events_csv"),
                "stage_a_stability": inputs.get("stage_a_stability_csv"),
                "stage_b_selection": inputs.get("stage_b_selection_csv"),
                "policy_summary": inputs.get("policy_summary_csv"),
                "capacity_prune_events": inputs.get("capacity_prune_events_csv"),
            }
            
            for sheet_name in sheets_to_create:
                if sheet_name == "summary":
                    continue
                
                csv_path = sheet_mapping.get(sheet_name)
                if csv_path is None:
                    continue
                
                rows = _read_csv_to_rows(csv_path)
                if rows is None:
                    sheets[sheet_name] = pd.DataFrame([{"status": "missing"}])
                    continue
                
                if not rows:
                    sheets[sheet_name] = pd.DataFrame([{"status": "empty"}])
                    continue
                
                # Читаем через pandas для совместимости
                try:
                    df = pd.read_csv(csv_path)
                    sheets[sheet_name] = df
                except Exception as e:
                    warnings.warn(
                        f"Failed to read {csv_path} with pandas: {e}. Creating missing sheet.",
                        UserWarning,
                        stacklevel=2
                    )
                    sheets[sheet_name] = pd.DataFrame([{"status": "error"}])
            
            # Используем pandas ExcelWriter напрямую
            with pd.ExcelWriter(output_path, engine="xlsxwriter") as writer:
                for sheet_name, df in sheets.items():
                    sheet_name_limited = sheet_name[:31]
                    df.to_excel(writer, sheet_name=sheet_name_limited, index=False)
            
            print(f"📦 Saved report pack to {output_path}")
            return output_path
            
    except Exception as e:
        warnings.warn(
            f"Failed to create report_pack.xlsx: {e}. Continuing without XLSX export.",
            UserWarning,
            stacklevel=2
        )
        return None

